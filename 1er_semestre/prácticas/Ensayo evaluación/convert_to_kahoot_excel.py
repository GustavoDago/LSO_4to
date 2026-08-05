import csv
from openpyxl import load_workbook

# Paths
template_path = r'F:\Mochila\Antigravity\NotebookLm\prácticas\practica_gamificada\KahootQuizTemplate.xlsx'
csv_path = r'F:\Mochila\Antigravity\NotebookLm\prácticas\Ensayo evaluación\kahoot_LSO_4to_anno.csv'
output_path = r'F:\Mochila\Antigravity\NotebookLm\prácticas\Ensayo evaluación\Kahoot_LSO_4to_anno_Final.xlsx'

# Load workbook
wb = load_workbook(template_path)
ws = wb.active

# Clear existing example data starting from row 9
for row in range(9, ws.max_row + 1):
    for col in range(1, 9):
        ws.cell(row=row, column=col).value = None

# Read CSV and populate Excel
with open(csv_path, 'r', encoding='utf-8') as f:
    reader = csv.reader(f)
    next(reader) # Skip CSV header
    
    row_idx = 9
    for i, row in enumerate(reader):
        question = row[0]
        ans1 = row[1]
        ans2 = row[2]
        ans3 = row[3]
        ans4 = row[4]
        time_limit = row[5]
        correct_ans = row[6]
        
        ws.cell(row=row_idx, column=1).value = i + 1
        ws.cell(row=row_idx, column=2).value = question
        ws.cell(row=row_idx, column=3).value = ans1
        ws.cell(row=row_idx, column=4).value = ans2
        ws.cell(row=row_idx, column=5).value = ans3
        ws.cell(row=row_idx, column=6).value = ans4
        ws.cell(row=row_idx, column=7).value = int(time_limit)
        ws.cell(row=row_idx, column=8).value = int(correct_ans)
        
        row_idx += 1

# Save output
wb.save(output_path)
print("Successfully generated Excel file at:", output_path)
